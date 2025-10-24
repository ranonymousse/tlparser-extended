import json
import openpyxl
import os
import pandas as pd
from contextlib import nullcontext
from datetime import datetime
from typing import Callable, Sequence, Tuple

import click

from tlparser.config import Configuration
from tlparser.stats import Stats
from tlparser.stats_ext import SpotAnalyzer


class Utils:

    def __init__(self, config: Configuration):
        self.config = config
        self.warnings: list[str] = []
        self.spot_issues: list[tuple[str, list[str]]] = []

    def read_formulas_from_json(
        self,
        *,
        extended: bool = False,
        verbose: bool = False,
        progress_factory: Callable[[int], object] | None = None,
    ):
        self.warnings.clear()
        self.spot_issues = []

        spot_analyzer = self._create_spot_analyzer(extended, verbose)

        with open(self.config.file_data_in, "r") as file:
            data = json.load(file)
        parsed_formulas = []

        ids = [item["id"] for item in data]
        if len(ids) != len(set(ids)):
            raise ValueError("Duplicate IDs found, abort...")
        progress_cm = nullcontext()
        if progress_factory is not None:
            total = sum(
                len(entry.get("logics", []))
                for entry in data
                if entry.get("status") in self.config.only_with_status
            )
            progress_cm = progress_factory(total)

        with progress_cm as progress:
            for entry in data:
                if entry["status"] in self.config.only_with_status:
                    for logic in entry["logics"]:
                        s = Stats(
                            formula_str=logic["f_code"],
                            req_text=entry["text"],
                            extended=extended,
                            spot_analyzer=spot_analyzer,
                            spot_verbose=verbose,
                        )
                        parsed_formulas.append(
                            {
                                "id": entry["id"],
                                "text": entry["text"],
                                "type": logic["type"],
                                "translation": logic["translation"],
                                "reasoning": logic["reasoning"],
                                "stats": s.get_stats(),
                            }
                        )
                        if progress is not None:
                            progress.update(1)
        if spot_analyzer is not None:
            self.warnings.extend(spot_analyzer.diagnostics)
            self.spot_issues.extend(spot_analyzer.issue_entries())
        return parsed_formulas

    def save_spot_issue_report(self, path: str) -> None:
        if not self.spot_issues:
            return
        lines = ["# Spot Analysis Issues", ""]
        for formula, problems in self.spot_issues:
            lines.extend(["## Formula", "```", formula, "```", "", "### Issues"])
            for issue in problems:
                lines.append(f"- {issue}")
            lines.append("")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines).rstrip() + "\n")

    def analyze_single_formula(
        self,
        formula: str,
        *,
        extended: bool = False,
        requirement_text: str | None = None,
        verbose: bool = False,
    ) -> Stats:
        self.warnings.clear()
        self.spot_issues = []

        spot_analyzer = self._create_spot_analyzer(extended, verbose)

        stats = Stats(
            formula_str=formula,
            req_text=requirement_text,
            extended=extended,
            spot_analyzer=spot_analyzer,
            spot_verbose=verbose,
        )

        if spot_analyzer is not None:
            self.warnings.extend(spot_analyzer.diagnostics)
            self.spot_issues.extend(spot_analyzer.issue_entries())

        return stats

    def _create_spot_analyzer(
        self, extended: bool, verbose: bool
    ) -> SpotAnalyzer | None:
        if not extended:
            return None
        try:
            return SpotAnalyzer(verbose=verbose)
        except Exception as exc:
            self.warnings.append(
                f"[tlparser] Spot analyzer initialization failed: {exc}"
            )
            return None

    def echo_spot_summary(
        self,
        issues: Sequence[Tuple[str, Sequence[str]]],
        *,
        total: int | None = None,
    ) -> None:
        if not issues:
            return
        count = len(issues)
        header = (
            f"Spot analysis summary ({count} of {total})"
            if total is not None
            else "Spot analysis summary"
        )
        click.echo(f"\n{header}:", err=True)
        for formula, problems in issues:
            self._echo_labelled("Formula:", formula, fg="cyan")
            issues_text = "\n".join(problems)
            self._echo_labelled("Issue:", issues_text, fg="red")
            click.echo("", err=True)

    def _echo_labelled(self, label: str, value: str, *, fg: str) -> None:
        styled_label = click.style(label, fg=fg, bold=True)
        indent = " " * len(label)
        lines = value.splitlines() or [""]
        first, *rest = lines
        click.echo(f"{styled_label} {first}", err=True)
        for line in rest:
            click.echo(f"{indent} {line}", err=True)

    def write_to_excel(self, data):
        flattened_data = [self.flatten_dict(item) for item in data]

        # Derive and append translation class
        df = pd.DataFrame(flattened_data)
        type_order = self.config.logic_order
        df["type"] = pd.Categorical(df["type"], categories=type_order, ordered=True)
        df_sorted = df.sort_values(by=["id", "type"])
        df["translationclass"] = df_sorted.groupby("id")["translation"].transform(
            lambda x: "".join([v[0] for v in x])
        )

        flattened_data = df.to_dict(orient="records")
        headers = {key for item in flattened_data for key in item.keys()}

        # Sort headers according to predefined order, with any extra headers at the end
        include_extended = any("stats.spot" in entry for entry in flattened_data)
        predefined_order = Utils.get_column_order(extended=include_extended)
        headers = [header for header in predefined_order if header in headers] + [
            header for header in headers if header not in predefined_order
        ]

        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write the headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Write the data to the sheet
        for row, item in enumerate(flattened_data, start=2):
            for col, header in enumerate(headers, start=1):
                value = item.get(header)
                if (
                    value is None
                    or value == ""
                    or (isinstance(value, set) and len(value) == 0)
                ):
                    sheet.cell(row=row, column=col, value=None)
                elif isinstance(value, int) or isinstance(value, float):
                    sheet.cell(row=row, column=col, value=value)
                elif isinstance(value, set):
                    sheet.cell(row=row, column=col, value=" | ".join(value))
                else:
                    sheet.cell(row=row, column=col, value=str(value))

        # Save the workbook to a file
        os.makedirs(self.config.folder_data_out, exist_ok=True)
        prefix = self.extract_filename_without_suffix(self.config.file_data_in)
        out = os.path.join(
            self.config.folder_data_out, f"{prefix}_{self.get_unique_filename()}.xlsx"
        )
        workbook.save(out)
        return out

    @staticmethod
    def flatten_dict(d, parent_key="", sep="."):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(Utils.flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

    @staticmethod
    def get_unique_filename():
        now = datetime.now()
        return now.strftime("%y%m%d%H%M%S")

    @staticmethod
    def extract_filename_without_suffix(file_path):
        base_name = os.path.basename(file_path)
        return os.path.splitext(base_name)[0]

    @staticmethod
    def get_latest_excel(folder):
        excel_files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
        if not excel_files:
            return ""

        # Sort files by modification time
        latest_file = max(
            excel_files, key=lambda f: os.path.getmtime(os.path.join(folder, f))
        )
        file = os.path.join(folder, latest_file)
        return file

    @staticmethod
    def rotate_palette_map(
        palette,
        types,
        *,
        index: int = 0,
        default_color: str = "#808080",
    ):
        """Return a dict mapping of types to colors, rotated by index.

        - palette: mapping type -> color (preferred). If not a dict, attempts to
          treat it as a sequence of colors aligned to the provided types.
        - types: ordered sequence of type labels to include.
        - index: rotation amount (wraps; negative allowed). 0 keeps original.
        - default_color: color used when a type is missing from palette.
        """
        types = list(types)
        if not types:
            return {}

        # Build color list aligned to the given types
        if isinstance(palette, dict):
            colors = [palette.get(t, default_color) for t in types]
        else:
            # Fallback: assume palette is a sequence of colors
            palette_seq = list(palette or [])
            if not palette_seq:
                palette_seq = [default_color] * len(types)
            # Repeat or trim to match types length
            times = (len(types) + len(palette_seq) - 1) // len(palette_seq)
            colors = (palette_seq * times)[: len(types)]

        if len(colors) > 0 and isinstance(index, int):
            k = index % len(colors)
            if k:
                colors = colors[k:] + colors[:k]

        return dict(zip(types, colors))

    @staticmethod
    def lighten_color(hex_color, opacity=0.6):
        hex_color = hex_color.lstrip("#")
        r, g, b = (
            int(hex_color[0:2], 16),
            int(hex_color[2:4], 16),
            int(hex_color[4:6], 16),
        )
        white = (255, 255, 255)
        new_r = int((1 - opacity) * white[0] + opacity * r)
        new_g = int((1 - opacity) * white[1] + opacity * g)
        new_b = int((1 - opacity) * white[2] + opacity * b)
        return f"#{new_r:02x}{new_g:02x}{new_b:02x}"

    @staticmethod
    def get_column_order(extended: bool = False):
        base_columns = [
            "id",
            "text",
            "type",
            "reasoning",
            "translation",
            "translationclass",
            "stats.req_len",
            "stats.req_sentence_count",
            "stats.req_word_count",
            "stats.formula_raw",
            "stats.formula_parsable",
            "stats.formula_parsed",
            "stats.asth",
            "stats.ap",
            "stats.cops.eq",
            "stats.cops.geq",
            "stats.cops.gt",
            "stats.cops.leq",
            "stats.cops.lt",
            "stats.cops.neq",
            "stats.lops.and",
            "stats.lops.impl",
            "stats.lops.not",
            "stats.lops.or",
            "stats.tops.A",
            "stats.tops.E",
            "stats.tops.F",
            "stats.tops.G",
            "stats.tops.R",
            "stats.tops.U",
            "stats.tops.X",
            "stats.agg.aps",
            "stats.agg.cops",
            "stats.agg.lops",
            "stats.agg.tops",
            "stats.entropy.lops",
            "stats.entropy.tops",
            "stats.entropy.lops_tops",
        ]
        if not extended:
            return base_columns

        spot_columns = [
            "stats.spot.formula",
            "stats.spot.spot_formula",
            "stats.spot.syntactic_safety",
            "stats.spot.is_stutter_invariant_formula",
            "stats.spot.manna_pnueli_class",
            "stats.spot.tgba_analysis.state_count",
            "stats.spot.tgba_analysis.transition_count",
            "stats.spot.tgba_analysis.is_complete",
            "stats.spot.tgba_analysis.is_deterministic",
            "stats.spot.tgba_analysis.acceptance_sets",
            "stats.spot.tgba_analysis.is_stutter_invariant",
            "stats.spot.buchi_analysis.state_count",
            "stats.spot.buchi_analysis.transition_count",
            "stats.spot.buchi_analysis.is_complete",
            "stats.spot.buchi_analysis.is_deterministic",
            "stats.spot.buchi_analysis.acceptance_sets",
            "stats.spot.buchi_analysis.is_stutter_invariant",
            "stats.spot.deterministic_attempt.success",
            "stats.spot.deterministic_attempt.error",
            "stats.spot.deterministic_attempt.automaton_analysis.state_count",
            "stats.spot.deterministic_attempt.automaton_analysis.transition_count",
            "stats.spot.deterministic_attempt.automaton_analysis.is_complete",
            "stats.spot.deterministic_attempt.automaton_analysis.is_deterministic",
            "stats.spot.deterministic_attempt.automaton_analysis.acceptance_sets",
            "stats.spot.deterministic_attempt.automaton_analysis.is_stutter_invariant",
        ]
        return base_columns + spot_columns
